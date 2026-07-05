# -*- coding: utf-8 -*-
"""
NHAP HANG LOAT BAI DANG cho DHB Reup Facebook
==============================================
Cong cu NGOAI, KHONG dung toi file .exe/.dll cua tool.
No chi ghi du lieu vao file du an .dhb (SQLite) trong thu muc Data/Project,
giong het khi ban tu them tung bai qua GUI roi luu du an.

Cach hoat dong:
  1. Doc file Excel  (mau_bai_dang.xlsx)
  2. Lay "khuon" 1 bai truc tiep tu du lieu HIEN TAI cua ban (de luon dung phien ban tool)
  3. Resolve "Trang dang" -> tai khoan/token nho bang AccountInfo trong data.sqlite
  4. Tao ra file du an  NhapHangLoat.dhb  chua toan bo bai (Status = "Not start")
  -> Mo tool, Load du an "NhapHangLoat", bam chay.

An toan:
  - Tu dong kiem tra tool CO DANG CHAY khong (bat buoc tat truoc khi ghi).
  - Tu dong backup file du an cu vao thu muc backups/ truoc khi ghi.
  - Chi ghi vao file du an rieng (NhapHangLoat.dhb), KHONG dong vao du an co san.
"""

import os, sys, re, json, base64, sqlite3, shutil, random, argparse, subprocess, unicodedata, warnings
from datetime import datetime, timezone, timedelta

warnings.filterwarnings("ignore", message="Data Validation extension is not supported")

VERSION = "1.1"   # tang moi lan cap nhat -> de biet 2 may co cung ban khong

# ================== CAU HINH ==================
HERE      = os.path.dirname(os.path.abspath(__file__))
EXCEL     = os.path.join(HERE, "mau_bai_dang.xlsx")
BACKUPS   = os.path.join(HERE, "backups")

def _find_tool_data():
    """Tu tim thu muc 'Data' cua DHB Reup -> chuyen sang may khac khong phai sua code.
       Uu tien: file duong_dan_tool.txt -> bien moi truong -> cac vi tri pho bien."""
    def ok(p):
        return bool(p) and os.path.exists(os.path.join(p, "data.sqlite"))
    # 0) override RIENG TUNG MAY: file trong thu muc User (KHONG bi dong bo qua cloud).
    #    Dung khi de code trong OneDrive/Google Drive ma moi may cai DHB o cho khac nhau.
    hp = os.path.join(os.path.expanduser("~"), "DHB_duong_dan_tool.txt")
    if os.path.exists(hp):
        try:
            p = open(hp, encoding="utf-8-sig").read().strip().strip('"')
            if ok(p): return p
        except Exception:
            pass
    # 1) file 'duong_dan_tool.txt' canh script (nguoi dung tu dan duong dan neu can)
    cfg = os.path.join(HERE, "duong_dan_tool.txt")
    if os.path.exists(cfg):
        try:
            p = open(cfg, encoding="utf-8-sig").read().strip().strip('"')
            if ok(p): return p
        except Exception:
            pass
    # 2) bien moi truong
    env = os.environ.get("DHB_TOOL_DATA", "").strip().strip('"')
    if ok(env): return env
    # 3) cac vi tri ung vien (uu tien: ngay canh thu muc cong cu nay)
    cands = [
        os.path.join(HERE, "..", "DHB Reup Facebook", "DHB Reup Facebook", "Data"),
        os.path.join(os.path.expanduser("~"), "Documents", "09_TOOLS_INSTALLERS", "DHB_Tools",
                     "DHB Reup Facebook", "DHB Reup Facebook", "Data"),
        os.path.join(os.path.expanduser("~"), "Documents", "DHB Reup Facebook",
                     "DHB Reup Facebook", "Data"),
        os.path.join(os.path.expanduser("~"), "Desktop", "DHB Reup Facebook",
                     "DHB Reup Facebook", "Data"),
    ]
    for c in cands:
        if ok(c): return os.path.abspath(c)
    return os.path.abspath(cands[0])   # khong tim thay -> check_paths() se bao loi ro rang

TOOL_DATA = _find_tool_data()

DATA_SQLITE   = os.path.join(TOOL_DATA, "data.sqlite")
OPTION_SQLITE = os.path.join(TOOL_DATA, "DataOptionUpload.sqlite")
PROJECT_DIR   = os.path.join(TOOL_DATA, "Project")

VN_TZ = timezone(timedelta(hours=7))     # gio Viet Nam +07:00
SCHEDULE_SUFFIX = "$99"                    # hau to tool luon gan vao Schedule

# Ten cot trong Excel (khong phan biet hoa thuong / khoang trang thua)
COL = {
    "project": "Du an",                # tuy chon: ten du an cho tung bai (de trong = du an mac dinh)
    "url":     "Nguon (URL)",
    "source_type": "Loai nguon",       # tuy chon: Direct Link / Youtube / Facebook / File... (de trong = tu nhan dang)
    "caption": "Caption / Mo ta",
    "page":    "Trang dang",
    "account": "Tai khoan",            # tuy chon, chi can khi 1 ten trang trung o nhieu tai khoan
    "type":    "Dang dang",            # Reels / Bai viet / Story
    "sched":   "Lich dang (YYYY-MM-DD HH:MM)",
    "group":   "Nhom tuy chon (link Shopee...)",  # tuy chon
    "title":   "Tieu de",              # tuy chon
    "link":    "Link",                 # tuy chon
    "status":  "Trang thai",           # chi dung khi SUA project: 'Not start' = se chay lai
}

UPLOAD_MAP = {   # gia tri Excel -> gia tri tool (theo dung chuoi tool dung)
    "reels": "Reels", "reel": "Reels",
    "bai viet": "Post", "post": "Post", "posts": "Post", "feed": "Post",
    "story": "Story",
}
UPLOAD_REV = {"Reels": "Reels", "Post": "Bai viet", "Story": "Story"}   # tool -> Excel (khi xuat)

# Cot AN luu nguyen ban goc cua tung bai (de SUA khong mat du lieu). Dung KHONG sua/xoa.
GOC_HEADER = "__goc__ (KHONG sua/xoa cot nay)"

MEDIA_EXT = (".mp4", ".mov", ".mkv", ".avi", ".webm", ".flv", ".m4v",
             ".wmv", ".ts", ".mpg", ".mpeg", ".3gp")

# Loai nguon da duoc kiem chung/khop du lieu that. Loai ngoai danh sach & chua co khuon -> canh bao.
VERIFIED_TYPES = {"Direct Link", "File", "Youtube", "Facebook"}

# Loai nguon (Itype) - dung DUNG ten hien thi trong hop "Chon dau vao" cua tool
SRC_CANON = {
    "direct link": "Direct Link", "directlink": "Direct Link", "direct": "Direct Link", "link": "Direct Link",
    "file": "File",
    "facebook": "Facebook", "fb": "Facebook",
    "youtube": "Youtube", "yt": "Youtube",
    "youtube-dl": "Youtube-DL", "youtubedl": "Youtube-DL", "ytdl": "Youtube-DL",
    "dailymotion": "Dailymotion",
}
# Chi File (duong dan may) moi nhet duong dan vao Source[0].Url (giong bai File that cua ban).
# Direct Link / Youtube / Facebook... de Source rong - tool tai tu Url (dung nhu GUI tool tao ra).
INLINE_SOURCE = {"File"}

# ================== TIEN ICH ==================
def norm(s):
    return re.sub(r"\s+", " ", str(s or "").strip()).lower()

def strip_accents(s):
    """Bo dau tieng Viet de so khop loai bai (Reels/Bai viet/Story)."""
    s = unicodedata.normalize("NFD", str(s or ""))
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    return s.replace("đ", "d").replace("Đ", "D").strip().lower()

class DhbError(Exception):
    """Loi 'mem' - hien thi cho nguoi dung (console: in ra; GUI: hop thoai)."""
    pass

def die(msg, code=1):
    raise DhbError(msg)

def info(msg): print("  " + msg)

def ensure_tool_closed():
    """Kiem tra DHB Reup.exe co dang chay khong."""
    try:
        out = subprocess.run(["tasklist", "/FI", "IMAGENAME eq DHB Reup.exe"],
                             capture_output=True, text=True, timeout=15).stdout
    except Exception:
        out = ""
    if "DHB Reup.exe" in out:
        die("Tool DHB Reup DANG CHAY. Hay TAT tool hoan toan roi chay lai.\n"
            "      (Ghi du lieu luc tool dang mo co the lam HONG database.)")

def check_paths():
    if not os.path.exists(os.path.join(TOOL_DATA, "data.sqlite")):
        die("Khong tim thay thu muc 'Data' cua tool DHB Reup.\n"
            "      Dang thu: " + TOOL_DATA + "\n\n"
            "      CACH SUA (nhat la khi vua chuyen sang may moi):\n"
            "      Tao 1 file ten 'duong_dan_tool.txt' trong thu muc _NhapHangLoat,\n"
            "      dan vao duong dan thu muc Data cua DHB Reup roi luu lai. Vi du noi dung:\n"
            "      C:\\Users\\<TenBan>\\...\\DHB Reup Facebook\\DHB Reup Facebook\\Data")
    for p in (DATA_SQLITE, PROJECT_DIR):
        if not os.path.exists(p):
            die("Thieu file/thu muc trong Data: " + p)

# ================== TAI KHOAN / TRANG ==================
def load_accounts():
    """Tra ve (pages, profiles) tu bang AccountInfo.
       pages: dict  ten_trang_chuan -> list cac dict Account{Type,ID,Name,IDProfile,NameAccount}
       profiles: dict  ten_profile_chuan -> Account dict
    """
    con = sqlite3.connect(DATA_SQLITE); c = con.cursor()
    rows = c.execute("SELECT ID, IDProfile, Name, Type FROM AccountInfo").fetchall()
    con.close()
    prof_name = {}   # IDProfile -> ten hien thi cua tai khoan
    for _id, idp, name, typ in rows:
        if typ == "Profile":
            prof_name[str(idp)] = name
    pages, profiles = {}, {}
    for _id, idp, name, typ in rows:
        acc = {"Type": typ, "ID": str(_id), "Name": name,
               "IDProfile": str(idp), "NameAccount": prof_name.get(str(idp), name)}
        if typ == "Page":
            pages.setdefault(norm(name), []).append(acc)
        elif typ == "Profile":
            profiles[norm(name)] = acc
    return pages, profiles

def load_fangroups():
    """Tra ve list ten cac NHOM fanpage (bang 'Group' trong data.sqlite).
       Dang len 1 nhom = dang len tat ca page trong nhom (tool tu tra luc chay)."""
    try:
        con = sqlite3.connect(f"file:{DATA_SQLITE}?mode=ro", uri=True)
        names = [r[0] for r in con.execute('SELECT NameGroup FROM "Group"')]
        con.close()
        return names
    except Exception:
        return []

def load_fangroup_members():
    """dict {ten_nhom: [ten_page, ...]} tu bang GroupInfo."""
    out = {}
    try:
        con = sqlite3.connect(f"file:{DATA_SQLITE}?mode=ro", uri=True)
        for ng, name in con.execute("SELECT NameGroup, Name FROM GroupInfo"):
            out.setdefault(ng, []).append(name)
        con.close()
    except Exception:
        pass
    return out

GROUP_RE = re.compile(r'(?i)^\s*(?:group|nhom|nhóm)\s*:\s*(.+)$')

def resolve_account(page_text, account_text, pages, profiles, fangroups=()):
    raw = str(page_text or "").strip()
    def as_group(name):
        return {"Type": "GROUP", "ID": None, "Name": name, "IDProfile": None, "NameAccount": None}

    # 1) Ghi ro dang nhom: "GROUP: xxx" / "Nhom: xxx"
    m = GROUP_RE.match(raw)
    if m:
        want = norm(m.group(1))
        for g in fangroups:
            if norm(g) == want:
                return as_group(g)
        raise ValueError(f'Khong tim thay NHOM "{m.group(1).strip()}". '
                         f'Xem sheet "DanhSachTrang" (cac dong Loai = GROUP).')

    key = norm(raw)
    # 2) Trang don le
    cands = pages.get(key)
    if cands:
        if len(cands) == 1:
            return cands[0]
        if account_text:
            akey = norm(account_text)
            for a in cands:
                if norm(a["NameAccount"]) == akey:
                    return a
        raise ValueError(f'Ten trang "{page_text}" co o {len(cands)} tai khoan '
                         f'({", ".join(a["NameAccount"] for a in cands)}). '
                         f'Hay dien them cot "{COL["account"]}".')
    # 3) Ten nhom viet truc tiep (khong co tien to)
    for g in fangroups:
        if norm(g) == key:
            return as_group(g)
    # 4) Trang ca nhan (profile)
    if key in profiles:
        return profiles[key]
    raise ValueError(f'Khong tim thay trang/nhom/tai khoan "{page_text}" trong danh sach cua tool. '
                     f'Xem sheet "DanhSachTrang" de copy dung ten.')

def load_option_groups():
    """Tra ve dict  ten_nhom_chuan -> OptionGroup dict {ID,Name,Options[]}."""
    groups = {}
    if not os.path.exists(OPTION_SQLITE):
        return groups
    con = sqlite3.connect(OPTION_SQLITE); c = con.cursor()
    grp = {}
    for gid, data in c.execute("SELECT ID, OptionData FROM UploadGroups"):
        try: gj = json.loads(data)
        except Exception: continue
        grp[gid] = {"ID": gid, "Name": gj.get("Name", str(gid)), "Options": []}
    for oid, gidref, data in c.execute("SELECT ID, GroupID, OptionData FROM UploadOptions"):
        try: oj = json.loads(data)
        except Exception: continue
        if gidref in grp:
            grp[gidref]["Options"].append(oj)
    con.close()
    for g in grp.values():
        groups[norm(g["Name"])] = g
    return groups

# ================== KHUON BAI (hoc tu chinh du lieu cua ban) ==================
def harvest_posts():
    """Doc TAT CA bai tu moi du an .dhb hien co (chi doc, mo che do read-only)."""
    posts = []
    for fn in os.listdir(PROJECT_DIR):
        if not fn.lower().endswith(".dhb"): continue
        path = os.path.join(PROJECT_DIR, fn)
        if os.path.getsize(path) == 0: continue
        try:
            con = sqlite3.connect(f"file:{path}?mode=ro", uri=True)
            rows = con.execute("SELECT Post FROM Data").fetchall(); con.close()
        except Exception:
            continue
        for (b64,) in rows:
            try:
                posts.append(json.loads(base64.b64decode(b64).decode("utf-8")))
            except Exception:
                continue
    return posts

def build_templates(posts):
    """Tra ve (templates_by_Itype, generic).
       => moi LOAI NGUON lay khuon tu chinh bai cung loai ban tung tao (uu tien 'Not start'),
          nen dinh dang khop 100% voi phien ban tool cua ban. Loai chua tung dung -> khuon chung."""
    by_type, generic = {}, None
    for o in posts:
        st = o.get("Status")
        if generic is None or (st == "Not start" and generic.get("Status") != "Not start"):
            generic = o
        it = str(o.get("Itype"))
        cur = by_type.get(it)
        if cur is None or (st == "Not start" and cur.get("Status") != "Not start"):
            by_type[it] = o
    if generic is None:
        die("Chua co du an .dhb nao co bai de lam khuon. Hay tao 1 bai bat ky trong tool 1 lan roi thu lai.")
    return by_type, generic

# ================== TAO BAI ==================
def rand_idpost():
    return "".join(random.choice("0123456789ABCDEF") for _ in range(16))

def make_idpost(url, stype):
    # CHI Facebook moi lay ID so tu URL. Cac nguon khac (Direct Link/Youtube/File...)
    # dung ma hex ngau nhien - dung nhu tool tu sinh, tranh vo nham cum so trong link.
    if stype == "Facebook":
        for pat in (r"/reel/(\d+)", r"/videos/(\d+)", r"[?&]v=(\d+)",
                    r"/permalink/(\d+)", r"/posts/(\d+)"):
            m = re.search(pat, url or "")
            if m:
                return m.group(1)
    return rand_idpost()

def detect_source_type(url):
    """Tu nhan dang loai nguon tu URL/duong dan khi cot 'Loai nguon' de trong."""
    u = (url or "").strip(); ul = u.lower()
    if re.match(r"^[a-zA-Z]:[\\/]", u) or u.startswith("\\\\"):
        # duong dan may tinh: co duoi video -> File; khong co duoi (la thu muc) -> Folder
        return "File" if os.path.splitext(ul)[1] in MEDIA_EXT else "Folder"
    if "facebook.com" in ul or "fb.watch" in ul or "fb.com" in ul:
        return "Facebook"
    if "youtube.com" in ul or "youtu.be" in ul:
        return "Youtube"
    if "dailymotion.com" in ul or "dai.ly" in ul:
        return "Dailymotion"
    return "Direct Link"   # link http truc tiep (vd R2/CDN) -> Direct Link

def resolve_source_type(explicit, url):
    e = strip_accents(explicit)
    return SRC_CANON.get(e) or detect_source_type(url)

def parse_schedule(val):
    """Tra ve chuoi ISO +07:00, hoac None neu de trong."""
    if val is None or str(val).strip() == "":
        return None
    if isinstance(val, datetime):
        dt = val
    else:
        s = str(val).strip().replace("T", " ")
        dt = None
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%d/%m/%Y %H:%M",
                    "%d/%m/%Y %H:%M:%S", "%Y-%m-%d"):
            try: dt = datetime.strptime(s, fmt); break
            except ValueError: continue
        if dt is None:
            raise ValueError(f'Lich dang "{val}" khong doc duoc. Dung dang: 2026-07-05 08:30')
    return dt.replace(tzinfo=VN_TZ).isoformat()

def build_post(row, idx, pages, profiles, fangroups, groups, templates, generic):
    import copy
    url     = str(row.get("url", "") or "").strip()
    stype   = resolve_source_type(row.get("source_type", ""), url)

    # Chon KHUON:
    #  - Neu co ban goc an (cot __goc__, tuc dang SUA bai cu) -> dung chinh ban goc do (khong mat gi).
    #  - Neu khong (bai moi) -> dung khuon theo loai nguon (hoc tu du lieu cua ban).
    goc = row.get("_goc")
    if goc:
        try:
            p = json.loads(base64.b64decode(goc).decode("utf-8"))
        except Exception:
            p = copy.deepcopy(templates.get(stype) or generic); goc = None
    else:
        p = copy.deepcopy(templates.get(stype) or generic)

    caption = row.get("caption", "")
    caption = "" if caption is None else str(caption)
    page    = row.get("page", "")
    acc     = resolve_account(page, row.get("account", ""), pages, profiles, fangroups)

    up_raw  = strip_accents(row.get("type", "") or "reels")
    upload_to = UPLOAD_MAP.get(up_raw, "Reels")

    sched   = parse_schedule(row.get("sched"))

    # --- gan gia tri tung bai (phan noi dung/meta - luon cap nhat theo Excel) ---
    p["Url"]         = url
    p["Itype"]       = stype
    p["SourceType"]  = None
    p["Description"] = caption
    _t = str(row.get("title") or "").strip()
    _l = str(row.get("link") or "").strip()
    if not goc or _t: p["Title"] = _t or None   # SUA de trong -> giu Tieu de goc
    if not goc or _l: p["Link"]  = _l or None    # SUA de trong -> giu Link goc
    p["Account"]     = acc
    p["UploadTo"]    = upload_to
    p["Row"]         = idx

    # nhom tuy chon (link Shopee...) - tuy chon
    gtext = norm(row.get("group", ""))
    orig_og = p.get("OptionGroup") or {}
    orig_name = norm(orig_og.get("Name") or "")
    orig_is_none = (orig_og.get("ID") == 999) or (orig_name in ("none", ""))
    if goc and (gtext == orig_name or (not gtext and orig_is_none)):
        pass   # SUA bai cu ma KHONG doi nhom -> giu nguyen OptionGroup goc (khong dung lai => khong lech)
    elif gtext and gtext in groups:
        g = groups[gtext]
        p["OptionGroup"] = {"ID": g["ID"], "Name": g["Name"], "Options": g["Options"]}
    else:
        p["OptionGroup"] = {"ID": 999, "Name": "None", "Options": None}

    # lich / che do  (dang luon: Content=Public, Schedule="None" - dung nhu GUI tao)
    if sched:
        p["Content"]  = "Scheduled"
        p["Schedule"] = sched + SCHEDULE_SUFFIX
    else:
        p["Content"]  = "Public"
        p["Schedule"] = "None"

    # --- Trang thai chay ---
    # Bai moi, hoac cot "Trang thai" = 'Not start'/de trong  -> RESET de tool chay lai tu dau.
    # Nguoc lai (vd 'Hoan thanh')  -> GIU nguyen tien do/Source/IDPost cu (KHONG chay lai, tranh dang trung).
    want   = strip_accents(row.get("status", ""))
    rerun  = (not goc) or (want in ("", "not start", "notstart"))
    if rerun:
        p["IDPost"]        = make_idpost(url, stype)
        p["Source"]        = [{"Type": "video", "Url": url}] if stype in INLINE_SOURCE else []
        p["Status"]        = "Not start"
        p["Download"]      = 0
        p["Render"]        = 0
        p["Upload"]        = 0
        p["isRunning"]     = False
        p["isSuccess"]     = False
        p["LogError"]      = []
        p["ProcessIndex"]  = 0
        p["TempVideo"]     = None
        p["Output"]        = None
        p["InputVideo"]    = None
        p["OutputVideo"]   = None
        p["InputThumb"]    = None
        p["OutputThumb"]   = None
        p["Thumbnail"]     = None
        p["ThumbnailShow"] = None
        p["Duration"]      = "00:00:00"
        p["TotalDuration"] = "00:00:00"
    # neu khong rerun: giu nguyen Status/Download/Render/Upload/Source/IDPost tu ban goc.
    return p

# ================== GHI DU AN ==================
def get_create_sql():
    """Lay cau CREATE TABLE Data tu 1 du an co san (fallback neu khong co)."""
    for fn in os.listdir(PROJECT_DIR):
        if fn.lower().endswith(".dhb"):
            path = os.path.join(PROJECT_DIR, fn)
            if os.path.getsize(path) == 0: continue
            try:
                con = sqlite3.connect(path)
                r = con.execute("SELECT sql FROM sqlite_master WHERE type='table' AND name='Data'").fetchone()
                con.close()
                if r and r[0]: return r[0]
            except Exception:
                continue
    return 'CREATE TABLE "Data" ("Post" TEXT)'

def write_project(project_name, posts, append):
    target = os.path.join(PROJECT_DIR, project_name + ".dhb")
    # backup neu file da ton tai (ten LUON duy nhat - khong bao gio de len backup cu)
    if os.path.exists(target) and os.path.getsize(target) > 0:
        os.makedirs(BACKUPS, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        bak = os.path.join(BACKUPS, f"{project_name}_{stamp}.dhb")
        i = 1
        while os.path.exists(bak):
            bak = os.path.join(BACKUPS, f"{project_name}_{stamp}_{i}.dhb"); i += 1
        shutil.copy2(target, bak)
        info(f"Da backup du an cu -> {bak}")

    new_file = (not os.path.exists(target)) or os.path.getsize(target) == 0
    if new_file and os.path.exists(target):
        os.remove(target)

    con = sqlite3.connect(target); c = con.cursor()
    if new_file:
        c.execute(get_create_sql())
    else:
        # dam bao co bang Data
        if not c.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='Data'").fetchone():
            c.execute(get_create_sql())
        if not append:
            c.execute("DELETE FROM Data")   # xay lai tu dau tu Excel (idempotent)

    for p in posts:
        raw = json.dumps(p, ensure_ascii=False).encode("utf-8")
        c.execute('INSERT INTO Data ("Post") VALUES (?)', (base64.b64encode(raw).decode("ascii"),))
    con.commit(); con.close()
    return target

# ================== EXCEL ==================
def read_excel(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["BaiDang"] if "BaiDang" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [norm(h) for h in rows[0]]
    # ban do: khoa noi bo -> chi so cot
    idxmap = {}
    for key, label in COL.items():
        want = norm(label)
        for i, h in enumerate(headers):
            if h == want:
                idxmap[key] = i; break
    if "url" not in idxmap or "page" not in idxmap:
        die(f'File Excel thieu cot bat buoc "{COL["url"]}" hoac "{COL["page"]}". '
            f'Xoa file mau_bai_dang.xlsx roi chay lai de tao lai mau chuan.')
    # cot AN chua ban goc (khi SUA project)
    goc_i = None
    gn = norm(GOC_HEADER)
    for i, h in enumerate(headers):
        if h == gn:
            goc_i = i; break
    out = []
    for r in rows[1:]:
        rec = {}
        for key, i in idxmap.items():
            rec[key] = r[i] if i < len(r) else None
        if goc_i is not None and goc_i < len(r):
            rec["_goc"] = r[goc_i]
        if not (str(rec.get("url") or "").strip() or str(rec.get("caption") or "").strip()):
            continue   # bo dong trong
        out.append(rec)
    return out

def make_template_excel(path, pages, profiles, groups):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "BaiDang"
    headers = [COL["project"], COL["url"], COL["source_type"], COL["caption"], COL["page"],
               COL["account"], COL["type"], COL["sched"], COL["group"], COL["title"], COL["link"]]
    ws.append(headers)
    hdrfill = PatternFill("solid", fgColor="4472C4")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = hdrfill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    # 3 dong vi du (chi la mau, ban xoa/sua thoai mai). Cot 'Du an' -> gom bai vao du an cung ten.
    ex_page = next(iter(pages.values()))[0]["Name"] if pages else "Ten trang cua ban"
    ws.append(["DuAn_Reels", "https://pub-xxxx.r2.dev/tg/abc123_VI.mp4", "Direct Link",
               "Caption cua ban o day 😍\n#hashtag1 #hashtag2",
               ex_page, "", "Reels", "2026-07-05 08:30", "", "", ""])
    ws.append(["DuAn_Reels", "https://www.youtube.com/watch?v=xxxx", "",
               "Cung du an voi dong tren (DuAn_Reels).",
               ex_page, "", "Reels", "", "", "", ""])
    ws.append(["DuAn_Story", r"C:\Video\video1.mp4", "File",
               "De trong cot 'Du an' thi vao du an mac dinh.",
               ex_page, "", "Story", "", "", "", ""])
    widths = [16, 46, 14, 50, 26, 18, 12, 24, 26, 20, 24]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64+i)].width = w
    ws.freeze_panes = "B2"
    ranges = add_ref_sheets(wb, groups)
    from openpyxl.utils import get_column_letter
    order_keys = ["project", "url", "source_type", "caption", "page", "account",
                  "type", "sched", "group", "title", "link"]
    colmap = {k: get_column_letter(i + 1) for i, k in enumerate(order_keys)}
    add_dropdowns(ws, colmap, ranges)
    wb.save(path)

def add_ref_sheets(wb, groups):
    """Them sheet tra cuu (DanhSachTrang, NhomTuyChon) + sheet AN 'Lists' cho dropdown.
       Tra ve dict cac VUNG (range) de gan dropdown."""
    from openpyxl.styles import Font
    con = sqlite3.connect(DATA_SQLITE)
    rows = con.execute("SELECT ID,IDProfile,Name,Type FROM AccountInfo").fetchall()
    con.close()
    prof = {}
    for _id, idp, name, typ in rows:
        if typ == "Profile": prof[str(idp)] = name
    page_names = list(dict.fromkeys(name for _id, idp, name, typ in rows if typ == "Page"))
    acct_names = list(dict.fromkeys(prof.values()))
    fangroups  = load_fangroups()
    members    = load_fangroup_members()
    # trong dropdown "Trang dang": cac page don le + cac NHOM (ghi "GROUP: ten")
    page_choices = page_names + [f"GROUP: {g}" for g in fangroups]

    ws2 = wb.create_sheet("DanhSachTrang")
    ws2.append(["Ten trang / nhom (copy vao cot 'Trang dang')", "Tai khoan / gom page", "Loai", "ID"])
    for c in ws2[1]: c.font = Font(bold=True)
    for _id, idp, name, typ in rows:
        if typ == "Page":
            ws2.append([name, prof.get(str(idp), ""), typ, str(_id)])
    for g in fangroups:                     # cac nhom fanpage
        mem = members.get(g, [])
        ws2.append([f"GROUP: {g}", f"gom {len(mem)} page: " + "; ".join(mem), "GROUP", ""])
    ws2.column_dimensions["A"].width = 42
    ws2.column_dimensions["B"].width = 46
    ws2.column_dimensions["D"].width = 20

    # Sheet rieng: nhom gom nhung page nao
    if fangroups:
        ws4 = wb.create_sheet("DanhSachNhom")
        ws4.append(["Nhom (dung 'GROUP: ten' o cot Trang dang)", "So page", "Cac page trong nhom"])
        for c in ws4[1]: c.font = Font(bold=True)
        for g in fangroups:
            mem = members.get(g, [])
            ws4.append([f"GROUP: {g}", len(mem), "; ".join(mem)])
        ws4.column_dimensions["A"].width = 32
        ws4.column_dimensions["C"].width = 70

    ws3 = wb.create_sheet("NhomTuyChon")
    ws3.append(["Ten nhom (copy vao cot 'Nhom tuy chon')", "So option"])
    for c in ws3[1]: c.font = Font(bold=True)
    for g in groups.values():
        ws3.append([g["Name"], len(g["Options"])])
    ws3.column_dimensions["A"].width = 36

    # Sheet AN chua cac list cho dropdown (tham chieu VUNG -> tranh loi dau phay/cham phay theo ngon ngu may)
    wl = wb.create_sheet("Lists")
    data = {
        "A": ["Loai nguon", "Direct Link", "Youtube", "Facebook", "File", "Folder", "Youtube-DL", "Dailymotion"],
        "B": ["Dang dang", "Reels", "Bài viết", "Story"],
        "C": ["Trang thai", "Not start", "Hoàn thành"],
        "D": ["Trang"] + page_choices,
        "E": ["Tai khoan"] + acct_names,
    }
    for col, vals in data.items():
        for r, v in enumerate(vals, start=1):
            wl[f"{col}{r}"] = v
    wl.sheet_state = "hidden"

    def rng(col, n):
        return f"Lists!${col}$2:${col}${n + 1}" if n > 0 else None
    return {
        "src":     rng("A", 7),
        "up":      rng("B", 3),
        "status":  rng("C", 2),
        "page":    rng("D", len(page_choices)),
        "account": rng("E", len(acct_names)),
    }

def add_dropdowns(ws, colmap, ranges, max_row=2000):
    """Gan dropdown (list xo xuong) vao cac cot co tap gia tri co dinh."""
    from openpyxl.worksheet.datavalidation import DataValidation
    plan = [("source_type", "src"), ("type", "up"), ("status", "status"),
            ("page", "page"), ("account", "account")]
    for key, rkey in plan:
        col = colmap.get(key); r = ranges.get(rkey)
        if not col or not r:
            continue
        dv = DataValidation(type="list", formula1=r, allow_blank=True, showErrorMessage=False)
        dv.add(f"{col}2:{col}{max_row}")
        ws.add_data_validation(dv)

# ================== XUAT / NAP (sua project co san) ==================
def sched_to_display(s):
    s = str(s or "").strip()
    if not s or s == "None":
        return ""
    s = s.split("$")[0]
    try:
        return datetime.fromisoformat(s).strftime("%Y-%m-%d %H:%M")
    except Exception:
        return s

def post_to_row(o):
    acc = o.get("Account") or {}
    grp = o.get("OptionGroup") or {}
    gname = "" if (grp.get("ID") == 999 or str(grp.get("Name")) in ("None", "")) else str(grp.get("Name"))
    is_grp = str(acc.get("Type") or "").upper() == "GROUP"
    page_val = ("GROUP: " + (acc.get("Name") or "")) if is_grp else (acc.get("Name") or "")
    acct_val = "" if is_grp else (acc.get("NameAccount") or "")
    return {
        "url":         o.get("Url") or "",
        "source_type": o.get("Itype") or "",
        "caption":     o.get("Description") or "",
        "page":        page_val,
        "account":     acct_val,
        "type":        UPLOAD_REV.get(str(o.get("UploadTo")), str(o.get("UploadTo") or "Reels")),
        "sched":       sched_to_display(o.get("Schedule")),
        "group":       gname,
        "title":       o.get("Title") or "",
        "link":        o.get("Link") or "",
        "status":      o.get("Status") or "",
        "_goc":        base64.b64encode(json.dumps(o, ensure_ascii=False).encode("utf-8")).decode("ascii"),
    }

def read_project_posts(name):
    path = os.path.join(PROJECT_DIR, name + ".dhb")
    if not os.path.exists(path) or os.path.getsize(path) == 0:
        die(f'Khong tim thay du an "{name}" (hoac rong).')
    con = sqlite3.connect(f"file:{path}?mode=ro", uri=True)
    rows = con.execute("SELECT Post FROM Data").fetchall(); con.close()
    out = []
    for (b64,) in rows:
        try: out.append(json.loads(base64.b64decode(b64).decode("utf-8")))
        except Exception: continue
    return out

def sua_xlsx_path(name):
    return os.path.join(HERE, f"sua_{name}.xlsx")

def export_project(name, groups):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    posts = read_project_posts(name)
    order = ["url", "source_type", "caption", "page", "account", "type",
             "sched", "group", "title", "link", "status"]
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "BaiDang"
    ws.append([COL[k] for k in order] + [GOC_HEADER])
    hdrfill = PatternFill("solid", fgColor="4472C4")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF"); cell.fill = hdrfill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for o in posts:
        rec = post_to_row(o)
        ws.append([rec[k] for k in order] + [rec["_goc"]])
    widths = [46, 14, 50, 26, 18, 12, 22, 22, 18, 22, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64+i)].width = w
    goc_col = chr(64 + len(order) + 1)               # cot cuoi = ban goc
    ws.column_dimensions[goc_col].width = 12
    ws.column_dimensions[goc_col].hidden = True       # AN di, dung dung toi
    ws.freeze_panes = "A2"
    ranges = add_ref_sheets(wb, groups)
    from openpyxl.utils import get_column_letter
    colmap = {k: get_column_letter(i + 1) for i, k in enumerate(order)}
    add_dropdowns(ws, colmap, ranges)
    out = sua_xlsx_path(name)
    wb.save(out)
    return out, len(posts)

def apply_project(name, pages, profiles, groups, append=False):
    path_x = sua_xlsx_path(name)
    if not os.path.exists(path_x):
        die(f'Chua co file sua: {path_x}\n      Hay "Xuat du an" truoc de tao file nay.')
    posts_raw = read_excel(path_x)
    if not posts_raw:
        die("File sua khong co dong nao.")
    tmpl_by_type, generic = build_templates(harvest_posts())
    fangroups = load_fangroups()
    built, errors = [], []
    for i, row in enumerate(posts_raw):
        try:
            built.append(build_post(row, i, pages, profiles, fangroups, groups, tmpl_by_type, generic))
        except Exception as e:
            errors.append(f"  Dong {i+2}: {e}")
    if errors:
        print("\n[CANH BAO] Mot so dong loi, se BO QUA:")
        for e in errors: print(e)
    if not built:
        die("Khong co bai hop le de ghi.")
    target = write_project(name, built, append)
    n_run = sum(1 for p in built if p.get("Status") == "Not start")
    return target, len(built), n_run, len(errors)

# ================== DANH SACH DU AN ==================
def list_projects():
    items = []
    for fn in sorted(os.listdir(PROJECT_DIR)):
        if not fn.lower().endswith(".dhb"): continue
        path = os.path.join(PROJECT_DIR, fn)
        if os.path.getsize(path) == 0: continue
        try:
            con = sqlite3.connect(f"file:{path}?mode=ro", uri=True)
            n = con.execute("SELECT COUNT(*) FROM Data").fetchone()[0]; con.close()
        except Exception:
            n = 0
        items.append((fn[:-4], n))
    return items

# ================== CAC CHUC NANG ==================
def do_create(pages, profiles, groups, project=None, append=False):
    """Tao/ghi de du an tu file Excel mau_bai_dang.xlsx.
       - Moi bai vao du an ghi o cot 'Du an'; de trong -> du an mac dinh.
       - 1 file Excel co the tao NHIEU du an cung luc."""
    if not os.path.exists(EXCEL):
        make_template_excel(EXCEL, pages, profiles, groups)
        print("\n[OK] Da tao file mau (chua co san):")
        print("     " + EXCEL)
        print("  -> Mo file, dien cac bai vao sheet 'BaiDang' (cot 'Du an' = ten du an cho tung bai),")
        print("     luu lai, roi chay lai cong cu.")
        return
    menu_mode = project is None
    ensure_tool_closed()
    posts_raw = read_excel(EXCEL)
    if not posts_raw:
        die("File Excel chua co dong nao. Hay dien vao sheet 'BaiDang' roi chay lai.")
    # Chi hoi ten du an mac dinh khi CO dong de trong cot 'Du an'
    need_default = any(not str(r.get("project") or "").strip() for r in posts_raw)
    if project is None:
        if need_default:
            project = input("\nCo bai de trong cot 'Du an'. Ten du an mac dinh cho chung "
                            "(Enter = NhapHangLoat): ").strip() or "NhapHangLoat"
        else:
            project = "NhapHangLoat"   # moi bai da co ten du an rieng -> khong dung toi
    tmpl_by_type, generic = build_templates(harvest_posts())
    have = sorted(t for t in tmpl_by_type if t and t != "None")
    if have:
        info("Loai nguon da co khuon (khop tool): " + ", ".join(have))

    # gom bai theo du an (cot 'Du an'; de trong -> mac dinh)
    fangroups = load_fangroups()
    by_proj, errors = {}, []
    for i, row in enumerate(posts_raw):
        pj = (str(row.get("project") or "").strip() or project)
        try:
            lst = by_proj.setdefault(pj, [])
            lst.append(build_post(row, len(lst), pages, profiles, fangroups, groups, tmpl_by_type, generic))
        except Exception as e:
            errors.append(f"  Dong {i+2}: {e}")
    if errors:
        print("\n[CANH BAO] Mot so dong loi, se BO QUA:")
        for e in errors: print(e)
    if not by_proj:
        die("Khong co bai nao hop le de nhap.")

    all_built = [p for lst in by_proj.values() for p in lst]
    unverified = sorted({p["Itype"] for p in all_built
                         if p["Itype"] not in VERIFIED_TYPES and p["Itype"] not in tmpl_by_type})
    if unverified:
        print("\n[LUU Y] Loai nguon chua co khuon & chua kiem chung: " + ", ".join(unverified))
        print("        Nen tao 1 bai loai do trong tool + luu du an, roi chay lai (cong cu se tu hoc khuon).")

    # canh bao du an TRUNG TEN (da co san) -> cho chon Ghi de / Noi them / Huy
    def _exists(pj):
        p = os.path.join(PROJECT_DIR, pj + ".dhb")
        return os.path.exists(p) and os.path.getsize(p) > 0
    trung = [pj for pj in by_proj if _exists(pj)]
    if trung:
        print("\n[CHU Y] Cac du an sau DA CO SAN trong tool (trung ten): " + ", ".join(trung))
        if menu_mode:
            print("   g = GHI DE (xoa het bai cu, thay bang bai trong Excel; co tu backup)")
            print("   n = NOI THEM (giu bai cu, them bai moi vao cuoi)")
            print("   (phim khac = HUY, khong ghi gi)")
            ch = input("   Chon (g/n): ").strip().lower()
            if   ch == "g": append = False
            elif ch == "n": append = True
            else:
                print("Da HUY - khong ghi du an nao."); return
        else:
            print("   -> " + ("NOI THEM" if append else "GHI DE (tu backup ban cu)"))

    mode_txt = "noi them" if append else "ghi de, da backup"
    print(f"\n[HOAN TAT] Da nhap {len(all_built)} bai vao {len(by_proj)} du an:")
    for pj, lst in by_proj.items():
        write_project(pj, lst, append)
        tag = f"({mode_txt})" if pj in trung else "(moi)"
        print(f"   - {pj}: {len(lst)} bai  {tag}")
    print("  -> Mo tool -> load tung du an -> bam chay.")

def do_export(groups):
    """Xuat 1 du an ra Excel de sua."""
    items = list_projects()
    if not items:
        die("Chua co du an nao trong tool.")
    print("\nChon du an de XUAT ra Excel (de sua):")
    for i, (name, n) in enumerate(items, 1):
        print(f"   {i}. {name}   ({n} bai)")
    s = input("Nhap so du an: ").strip()
    if not s.isdigit() or not (1 <= int(s) <= len(items)):
        die("Khong chon du an hop le.")
    name = items[int(s) - 1][0]
    out, n = export_project(name, groups)
    print(f"\n[OK] Da xuat {n} bai ra: {out}")
    print("  -> Mo file do, sua o sheet 'BaiDang'.")
    print("     * Cot cuoi cung (dang AN) la ban goc tung bai - KHONG sua/xoa.")
    print("     * Muon 1 bai CHAY LAI: dat cot 'Trang thai' = Not start.")
    print("       (De nguyen 'Hoan thanh' thi bai do KHONG chay lai - tranh dang trung.)")
    print("     Sua xong -> luu lai -> chay cong cu -> chon muc 3 (Nap Excel da sua).")

def do_apply(pages, profiles, groups):
    """Nap file sua_*.xlsx da chinh, ghi de lai du an cung ten."""
    files = sorted(f for f in os.listdir(HERE)
                   if f.lower().startswith("sua_") and f.lower().endswith(".xlsx"))
    if not files:
        die("Chua co file sua_*.xlsx nao. Hay chon muc 2 (Xuat du an) truoc.")
    names = [f[4:-5] for f in files]
    print("\nChon file da sua de NAP (se ghi de lai du an cung ten):")
    for i, nm in enumerate(names, 1):
        print(f"   {i}. sua_{nm}.xlsx   ->  du an '{nm}'")
    s = input("Nhap so: ").strip()
    if not s.isdigit() or not (1 <= int(s) <= len(names)):
        die("Khong chon hop le.")
    name = names[int(s) - 1]
    print(f"\n! Se GHI DE toan bo du an '{name}' theo file sua_{name}.xlsx (tu backup ban cu).")
    if input("  Go 'y' roi Enter de tiep tuc: ").strip().lower() != "y":
        print("Da huy."); return
    ensure_tool_closed()
    target, n, n_run, nerr = apply_project(name, pages, profiles, groups, append=False)
    print(f"\n[HOAN TAT] Da ghi {n} bai vao du an '{name}'. Trong do {n_run} bai se chay (Not start).")
    print("     File: " + target)
    print(f"  -> Mo tool -> load lai du an '{name}' -> bam chay.")
    if nerr:
        print(f"  (Co {nerr} dong bi bo qua - xem canh bao phia tren.)")

def refresh_dropdowns_file(path, groups):
    """Lam moi danh sach trang/tai khoan (dropdown) trong 1 file Excel, GIU nguyen du lieu da dien."""
    import openpyxl
    from openpyxl.utils import get_column_letter
    wb = openpyxl.load_workbook(path)
    if "BaiDang" not in wb.sheetnames:
        return False
    ws = wb["BaiDang"]
    # xoa cac sheet phu + dropdown cu
    for s in ("DanhSachTrang", "NhomTuyChon", "Lists"):
        if s in wb.sheetnames:
            wb.remove(wb[s])
    ws.data_validations.dataValidation = []
    # dung lai colmap tu dong tieu de HIEN TAI (chiu duoc ca khi ban doi thu tu cot)
    label2key = {norm(v): k for k, v in COL.items()}
    colmap = {}
    for i, cell in enumerate(ws[1]):
        k = label2key.get(norm(cell.value))
        if k:
            colmap[k] = get_column_letter(i + 1)
    ranges = add_ref_sheets(wb, groups)
    add_dropdowns(ws, colmap, ranges)
    wb.save(path)
    return True

def do_refresh(groups):
    """Cap nhat danh sach trang/tai khoan cho mau_bai_dang.xlsx + tat ca sua_*.xlsx."""
    targets = []
    if os.path.exists(EXCEL):
        targets.append(EXCEL)
    for f in sorted(os.listdir(HERE)):
        if f.lower().startswith("sua_") and f.lower().endswith(".xlsx"):
            targets.append(os.path.join(HERE, f))
    if not targets:
        die("Chua co file Excel nao de cap nhat (mau_bai_dang.xlsx hoac sua_*.xlsx).")
    print("\nCap nhat danh sach (trang/tai khoan) - giu nguyen du lieu da dien:")
    ok = 0
    for p in targets:
        try:
            if refresh_dropdowns_file(p, groups):
                print(f"   [OK] {os.path.basename(p)}"); ok += 1
        except PermissionError:
            print(f"   [BO QUA] {os.path.basename(p)} - dang MO trong Excel, hay dong file roi thu lai.")
        except Exception as e:
            print(f"   [LOI] {os.path.basename(p)}: {e}")
    print(f"\nDa cap nhat {ok}/{len(targets)} file.")

# ================== MAIN ==================
def main():
    ap = argparse.ArgumentParser(description="Nhap / sua hang loat bai dang cho DHB Reup Facebook")
    ap.add_argument("--du-an", help="Tao/ghi de du an ten nay tu Excel mau")
    ap.add_argument("--noi-them", action="store_true", help="Noi them thay vi ghi de")
    ap.add_argument("--tao-mau", action="store_true", help="Tao lai file Excel mau roi thoat")
    ap.add_argument("--xuat", help="Xuat du an ten nay ra Excel de sua")
    ap.add_argument("--nap", help="Nap file sua_<ten>.xlsx, ghi de lai du an")
    ap.add_argument("--capnhat", action="store_true", help="Cap nhat danh sach trang/tai khoan trong cac file Excel")
    args = ap.parse_args()

    print("="*64)
    print("  NHAP / SUA HANG LOAT BAI DANG  ->  DHB Reup Facebook")
    print("="*64)
    check_paths()
    pages, profiles = load_accounts()
    groups = load_option_groups()
    info(f"Co {sum(len(v) for v in pages.values())} trang, "
         f"{len(profiles)} tai khoan, {len(groups)} nhom tuy chon.")

    # --- che do dong lenh (khong can menu) ---
    if args.tao_mau:
        make_template_excel(EXCEL, pages, profiles, groups)
        print("\n[OK] Da tao file mau: " + EXCEL); input("\nNhan Enter de dong..."); return
    if args.xuat:
        out, n = export_project(args.xuat, groups)
        print(f"\n[OK] Da xuat {n} bai -> {out}"); input("\nNhan Enter de dong..."); return
    if args.nap:
        ensure_tool_closed()
        target, n, n_run, nerr = apply_project(args.nap, pages, profiles, groups, args.noi_them)
        print(f"\n[HOAN TAT] {n} bai vao '{args.nap}' ({n_run} se chay)."); input("\nNhan Enter de dong..."); return
    if args.du_an:
        do_create(pages, profiles, groups, args.du_an, args.noi_them)
        input("\nNhan Enter de dong..."); return
    if args.capnhat:
        do_refresh(groups); input("\nNhan Enter de dong..."); return

    # --- menu (khi nhay dup file .bat) ---
    while True:
        print("\n---------------------------------------------")
        print("  1. Tao du an moi tu Excel (mau_bai_dang.xlsx)")
        print("  2. Xuat 1 du an ra Excel  de SUA")
        print("  3. Nap Excel da sua  ->  ghi de lai du an")
        print("  4. Cap nhat danh sach trang/tai khoan trong file Excel")
        print("  0. Thoat")
        c = input("Chon (0-4): ").strip()
        if   c == "1": do_create(pages, profiles, groups); break
        elif c == "2": do_export(groups); break
        elif c == "3": do_apply(pages, profiles, groups); break
        elif c == "4": do_refresh(groups); break
        elif c == "0": return
        else: print("  (Chon khong hop le, nhap 0-4)")
    input("\nNhan Enter de dong...")

if __name__ == "__main__":
    try:
        main()
    except DhbError as e:
        print("\n[LOI] " + str(e))
        input("\nNhan Enter de dong...")
    except SystemExit:
        raise
    except Exception as e:
        print("\n[LOI] Loi khong mong doi: " + repr(e))
        input("\nNhan Enter de dong...")
